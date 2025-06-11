from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from pydantic import BaseModel
import pandas as pd

# 1st --------------------------------------[Clean up markup - remove text with strike-through]-----------------------------------------------

workbook = load_workbook("sample_data/SRO658Markup.xlsx", rich_text=True)

# Always working with the first sheet
worksheet = workbook.worksheets[0]

# Iterate the row (skip the header row)
for row in worksheet.iter_rows(min_row=2): # 1 = header, 2 = first data row (skipping the header row)
    for cell in row:
        
        # If cell is empty or cell is a string without a strike-through, enter to continue (skip this cell - do nothing with it)
        if cell.value is None or isinstance(cell.value, str) and not cell.font.strike:
            continue # continue to next iteration - doesn't proceed with next if-condition
        
        # If entire cell is strike-through, clear the cell
        if isinstance(cell.value, str) and cell.font.strike:
            cell.value = None
            continue # continue to next interation - doesn't proceed with next if-condition
        
        # Otherwise, at run-level formatting
        if isinstance(cell.value, CellRichText):
            filtered_runs = []

            for run in cell.value:
                if isinstance(run, TextBlock) and run.font.strike is None:
                    filtered_runs.append(run)
                elif isinstance(run, str):
                    filtered_runs.append(run)

            cell.value = CellRichText(filtered_runs)

workbook.save("sample_data/output_markup.xlsx")

# # Read and load the markup into dataframe.
# df_markup = pd.read_excel("sample_data/SRO658Markup.xlsx", engine="openpyxl")
# # Read and load the export from DOORS into dataframe.
# df_doors = pd.read_excel("sample_data/SRO658_DOORS.xlsx", engine="openpyxl")

